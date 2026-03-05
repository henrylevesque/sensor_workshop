import time
import os
import math
from datetime import datetime
import board
import busio
import adafruit_ads1x15.ads1115 as ADS
from adafruit_ads1x15.analog_in import AnalogIn

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    print("openpyxl not found. Installing...")
    os.system("sudo python3 -m pip install --break-system-packages openpyxl")
    from openpyxl import Workbook, load_workbook

# This script reads sound level data from a microphone (MAX4466) connected to an ADS1115 ADC.
# Data is collected at set intervals for a set duration and saved to an Excel file.
# Data is appended (added) to the file after each reading, so no data is lost if power fails.
# Text that follows the '#' symbol is a comment and is not executed as part of the code.

# INSTALLATION INSTRUCTIONS (run on Raspberry Pi):
# Step 1: Enable I2C on the Raspberry Pi
#   sudo raspi-config → Navigate to: Interface Options → I2C → Enable
#   Then reboot: sudo reboot
# Step 2: Verify I2C is working
#   sudo i2cdetect -y 1 (you should see device address: 48)
# Step 3: Install openpyxl for Excel export
#   sudo python3 -m pip install --break-system-packages openpyxl
# Step 4: Install Adafruit ADS1x15 library
#   sudo python3 -m pip install --break-system-packages adafruit-circuitpython-ads1x15

# Wiring the Hardware
# MAX4466 Microphone to ADS1115 ADC to Raspberry Pi:
# MAX4466 VCC (labeled VCC on sensor) → ADS1115 VDD
# MAX4466 GND (labeled GND on sensor) → ADS1115 GND
# MAX4466 OUT (labeled OUT on sensor) → ADS1115 A0
#
# ADS1115 ADC to Raspberry Pi:
# ADS1115 VDD → 3.3V (Physical Pin 1)
# ADS1115 GND → GND (Physical Pin 6)
# ADS1115 SCL → GPIO 3 (Physical Pin 5)
# ADS1115 SDA → GPIO 2 (Physical Pin 3)
# ADS1115 ADDR → GND (Physical Pin 6)

SAMPLE_RATE = 500          # Mic sampling rate (Hz)
WINDOW_SIZE = 200          # Samples per dB reading
GAIN = 1                   # ADS gain (1 = ±4.096V)

# Setup ADS1115
def get_channel():
    """Initialize I2C and ADS1115 ADC"""
    i2c = busio.I2C(board.SCL, board.SDA)
    ads = ADS.ADS1115(i2c)
    ads.gain = GAIN
    return AnalogIn(ads, ADS.P0)

def read_db():
    """Read sound level in decibels"""
    chan = get_channel()
    samples = []

    for _ in range(WINDOW_SIZE):
        samples.append(chan.voltage)
        time.sleep(1 / SAMPLE_RATE)

    mean = sum(samples) / len(samples)
    centered = [s - mean for s in samples]
    rms = math.sqrt(sum(s**2 for s in centered) / len(centered))

    if rms > 0:
        db = 20 * math.log10(rms / 4.096)
    else:
        db = -100

    return db

# Function to create Excel file with headers (if it doesn't exist)
def init_excel_file(filename):
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.append(["Timestamp", "Sound Level (dB)"])
        wb.save(filename)
        print(f"Created new file: {filename}")

# Function to append a single row of data to Excel
def append_to_excel(filename, timestamp, db_level):
    wb = load_workbook(filename)
    ws = wb.active
    ws.append([timestamp, db_level])
    wb.save(filename)

# Function to collect data for a specified duration and interval
def collect_data(duration_minutes, interval_seconds, filename="sound_data.xlsx"):
    start_time = time.time()
    duration_seconds = duration_minutes * 60
    reading_count = 0
    
    # Create Excel file with headers
    init_excel_file(filename)
    
    print(f"Starting sound level recording for {duration_minutes} minutes (reading every {interval_seconds} seconds)...")
    print(f"Data will be saved to: {filename}")
    print("-" * 70)

    while time.time() - start_time < duration_seconds:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            db_level = read_db()
            reading_count += 1
            append_to_excel(filename, timestamp, db_level)
            print(f"[{reading_count}] {timestamp} | Sound Level: {db_level:.2f} dB")
        except Exception as e:
            print(f"[FAILED] {timestamp} | Sensor read error: {e}")
        
        time.sleep(interval_seconds)

    print("-" * 70)
    print(f"Recording complete! {reading_count} readings saved to {filename}")

# Example usage
if __name__ == "__main__":
    duration_minutes = 5      # Set the duration in minutes
    interval_seconds = 10     # Set the interval in seconds
    filename = "sound_data.xlsx"
    
    collect_data(duration_minutes, interval_seconds, filename)
