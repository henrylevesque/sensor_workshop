import time
import os
from datetime import datetime
try:
    from gpiozero import MotionSensor
except Exception:
    MotionSensor = None

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    print("openpyxl not found. Installing...")
    os.system("sudo python3 -m pip install --break-system-packages openpyxl")
    from openpyxl import Workbook, load_workbook

# This script reads data from a PIR motion sensor connected to a Raspberry Pi.
# Data is collected at set intervals for a set duration and saved to an Excel file.
# Data is appended (added) to the file after each reading, so no data is lost if power fails.
# Text that follows the '#' symbol is a comment and is not executed as part of the code.

# INSTALLATION INSTRUCTIONS (run on Raspberry Pi):
# Step 1: Install gpiozero for motion sensor control
#   sudo apt install python3-gpiozero
# Step 2: Install openpyxl for Excel export with append functionality
#   sudo python3 -m pip install --break-system-packages openpyxl

# Wiring the Hardware
# Connect PIR motion sensor to Raspberry Pi as follows:
# PIR VCC (labeled VCC on sensor) → 5V (Physical Pin 2)
# PIR GND (labeled GND on sensor) → GND (Physical Pin 6)
# PIR OUT (labeled OUT or SIGNAL on sensor) → GPIO 23 (Physical Pin 16)

# GPIO pin for the PIR motion sensor, change this if your sensor is connected to a different pin
PIN = 23

# Setup motion sensor
try:
    if MotionSensor is not None:
        sensor = MotionSensor(PIN)
    else:
        sensor = None
except Exception as e:
    print(f"Motion sensor init error: {e}")
    sensor = None

# Function to read PIR motion sensor data
def read_motion():
    if sensor is None:
        raise RuntimeError("Motion sensor not initialized. Install gpiozero with: sudo apt install python3-gpiozero")
    return sensor.is_active

# Function to create Excel file with headers (if it doesn't exist)
def init_excel_file(filename):
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.append(["Timestamp", "Motion Detected"])
        wb.save(filename)
        print(f"Created new file: {filename}")

# Function to append a single row of data to Excel
def append_to_excel(filename, timestamp, motion):
    wb = load_workbook(filename)
    ws = wb.active
    ws.append([timestamp, motion])
    wb.save(filename)

# Function to collect data for a specified duration and interval
def collect_data(duration_minutes, interval_seconds, filename="motion_data.xlsx"):
    start_time = time.time()
    duration_seconds = duration_minutes * 60
    reading_count = 0
    
    # Create Excel file with headers
    init_excel_file(filename)
    
    print(f"Starting motion detection for {duration_minutes} minutes (reading every {interval_seconds} seconds)...")
    print(f"Data will be saved to: {filename}")
    print("-" * 70)

    while time.time() - start_time < duration_seconds:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            motion = read_motion()
            reading_count += 1
            append_to_excel(filename, timestamp, motion)
            status = "MOTION DETECTED" if motion else "no motion"
            print(f"[{reading_count}] {timestamp} | {status}")
        except Exception as e:
            print(f"[FAILED] {timestamp} | Sensor read error: {e}")
        
        time.sleep(interval_seconds)

    print("-" * 70)
    print(f"Motion detection complete! {reading_count} readings saved to {filename}")

# Example usage
if __name__ == "__main__":
    duration_minutes = 5      # Set the duration in minutes
    interval_seconds = 10     # Set the interval in seconds
    filename = "motion_data.xlsx"
    
    collect_data(duration_minutes, interval_seconds, filename)
