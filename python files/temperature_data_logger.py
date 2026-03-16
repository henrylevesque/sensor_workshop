import time
import os
from datetime import datetime
try:
    import board
    import adafruit_dht
except Exception:
    board = None
    adafruit_dht = None

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    print("openpyxl not found. Installing...")
    os.system("sudo python3 -m pip install --break-system-packages openpyxl")
    from openpyxl import Workbook, load_workbook

# This script reads data from a DHT11 temperature sensor connected to a Raspberry Pi.
# Data is collected at set intervals for a set duration and saved to an Excel file.
# Data is appended (added) to the file after each reading, so no data is lost if power fails.
# Text that follows the '#' symbol is a comment and is not executed as part of the code.

# INSTALLATION INSTRUCTIONS (run on Raspberry Pi):
# Step 1: Install openpyxl for Excel export with append functionality
#   sudo python3 -m pip install --break-system-packages openpyxl
# Step 2: Install CircuitPython libraries for DHT support
#   sudo python3 -m pip install --break-system-packages adafruit-blinka
#   sudo python3 -m pip install --break-system-packages adafruit-circuitpython-dht

# IMPORTANT: Running this script with SSH
# If you start this script and then disconnect SSH, it will STOP.
# To keep this script running after you disconnect, use one of these methods:
#
# Method 1: Use nohup (recommended, simplest)
#   cd /home/pi
#   nohup python3 temperature_data_logger.py >/home/pi/temp_log.txt 2>&1 &
#
# Method 2: Start then disown
#   python3 temperature_data_logger.py >/home/pi/temp_log.txt 2>&1 &
#   jobs
#   disown -h %1    # prevents SIGHUP for job %1
#
# Both methods:
# - Redirect output to a file (>/path/to/log.txt) to keep from blocking
# - Allow you to safely disconnect SSH while the script runs
# - Can be stopped with: pkill -f temperature_data_logger.py
# - Data is appended to the Excel file throughout the run

# Wiring the Hardware
# Connect a Keyestudio DHT11 sensor to Raspberry Pi as follows:
# Keyestudio sensor pin labels (top to bottom on typical sensor):
#   S (Signal)  → GPIO 4 (Physical Pin 7)
#   V (VCC)     → 3.3V (Physical Pin 1)
#   G (GND)     → GND (Physical Pin 6)

# Create the DHT11 sensor object at GPIO 4 (board.D4 in CircuitPython notation)
try:
    if board is not None and adafruit_dht is not None:
        sensor = adafruit_dht.DHT11(board.D4)
    else:
        sensor = None
except Exception:
    sensor = None

# Function to read temperature and humidity data
def read_temperature():
    if sensor is None:
        raise RuntimeError(
            "DHT11 sensor not initialized. Install libraries with:\n"
            "  sudo python3 -m pip install --break-system-packages adafruit-blinka adafruit-circuitpython-dht"
        )
    
    try:
        temperature = sensor.temperature
        humidity = sensor.humidity
        return temperature, humidity
    except RuntimeError as e:
        print(f"Sensor read error (retry): {e}")
        return None, None

# Function to create Excel file with headers (if it doesn't exist)
def init_excel_file(filename):
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.append(["Timestamp", "Temperature (°C)", "Humidity (%)"])
        wb.save(filename)
        print(f"Created new file: {filename}")

# Function to append a single row of data to Excel
def append_to_excel(filename, timestamp, temperature, humidity):
    wb = load_workbook(filename)
    ws = wb.active
    ws.append([timestamp, temperature, humidity])
    wb.save(filename)

# Function to collect data for a specified duration and interval
def collect_data(duration_minutes, interval_seconds, filename="temperature_data.xlsx"):
    start_time = time.time()
    duration_seconds = duration_minutes * 60
    reading_count = 0
    
    # Create Excel file with headers
    init_excel_file(filename)
    
    print(f"Starting data collection for {duration_minutes} minutes (reading every {interval_seconds} seconds)...")
    print(f"Data will be saved to: {filename}")
    print("-" * 70)

    while time.time() - start_time < duration_seconds:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        temperature, humidity = read_temperature()
        
        if temperature is not None and humidity is not None:
            reading_count += 1
            append_to_excel(filename, timestamp, temperature, humidity)
            print(f"[{reading_count}] {timestamp} | Temp: {temperature:.1f}°C | Humidity: {humidity:.1f}%")
        else:
            print(f"[FAILED] {timestamp} | Sensor read error, will retry next interval")
        
        time.sleep(interval_seconds)

    print("-" * 70)
    print(f"Data collection complete! {reading_count} readings saved to {filename}")

# Example usage
if __name__ == "__main__":
    duration_minutes = 5      # Set the duration in minutes
    interval_seconds = 10     # Set the interval in seconds
    filename = "temperature_data.xlsx"
    
    collect_data(duration_minutes, interval_seconds, filename)
